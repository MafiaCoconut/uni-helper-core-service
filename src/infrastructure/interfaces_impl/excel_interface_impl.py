from openpyxl import Workbook

from application.interfaces.excel_interface import ExcelInterface
from infrastructure.config.logs_config import log_decorator


class ExcelInterfaceImpl(ExcelInterface):
    @staticmethod
    @log_decorator(print_args=False, print_kwargs=False)
    async def save_to_excel(headers: list, rows: list, path: str):
        wb = Workbook()
        ws = wb.active

        ws.append(headers)

        for row in rows:
            ws.append(row)

        for col_idx, column_header in enumerate(headers, start=1):
            max_length = max(len(str(cell)) for cell in ws[col_idx])
            adjusted_width = max_length + 2
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted_width

        wb.save(path)

    @staticmethod
    @log_decorator(print_args=False, print_kwargs=False)
    async def clear(path: str):
        wb = Workbook()
        ws = wb.active

        ws.delete_rows(1, ws.max_row)
        ws.delete_cols(1, ws.max_column)

        wb.save(path)

